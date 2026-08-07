import openpyxl

file_path = r'd:\AGRINAS PALMA NUSANTARA\AGRIPAM\DATA TK PANEN AGRIPAM.xlsx'
wb = openpyxl.load_workbook(file_path, data_only=True)

print("Sheet names:", wb.sheetnames)

for sheet_name in wb.sheetnames:
    sheet = wb[sheet_name]
    print(f"\n--- SHEET: {sheet_name} (Max row: {sheet.max_row}, Max col: {sheet.max_column}) ---")
    for r in range(1, 10):
        row_vals = [sheet.cell(r, c).value for c in range(1, sheet.max_column + 1)]
        if any(v is not None for v in row_vals):
            print(f"Row {r}:", row_vals)
