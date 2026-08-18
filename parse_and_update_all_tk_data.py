import openpyxl
import json
import os
import re

excel_path = r'd:\AGRINAS PALMA NUSANTARA\AGRIPAM\DATA TK PANEN AGRIPAM.xlsx'
excel_tambahan_path = r'd:\AGRINAS PALMA NUSANTARA\AGRIPAM\DATA TAMBAHAN PENUGASAN APN.xlsx'
json_path = r'd:\AGRINAS PALMA NUSANTARA\AGRIPAM\data_kebun_tk.json'
sql_path = r'd:\AGRINAS PALMA NUSANTARA\AGRIPAM\supabase\setup_kebun_tk_table.sql'
sql_root_path = r'd:\AGRINAS PALMA NUSANTARA\AGRIPAM\supabase_setup_data_kebun_tk.sql'

def normalize_region(r_raw):
    if not r_raw:
        return 'Aceh', 'Regional Aceh'
    s = str(r_raw).strip()
    raw = s
    s_upper = s.upper().replace('REGIONAL', '').replace('RO', '').strip()

    if 'SUMUT 1' in s_upper or 'SUMATERA UTARA 1' in s_upper: return 'Sumut 1', raw
    if 'SUMUT 2' in s_upper or 'SUMATERA UTARA 2' in s_upper or 'TORGANDA' in s_upper: return 'Sumut 2', raw
    if 'KALBAR 1' in s_upper or 'KALIMANTAN BARAT 1' in s_upper: return 'Kalbar 1', raw
    if 'KALBAR 2' in s_upper or 'KALIMANTAN BARAT 2' in s_upper: return 'Kalbar 2', raw
    if 'KALSEL 1' in s_upper or 'KALIMANTAN SELATAN 1' in s_upper: return 'Kalsel 1', raw
    if 'KALSEL 2' in s_upper or 'KALIMANTAN SELATAN 2' in s_upper: return 'Kalsel 2', raw
    if 'KALTARA' in s_upper or 'KALIMANTAN UTARA' in s_upper: return 'Kaltara', raw
    if 'KALTIM' in s_upper or 'KALIMANTAN TIMUR' in s_upper: return 'Kaltim', raw
    if 'KALTENG 1' in s_upper or 'KALIMANTAN TENGAH 1' in s_upper: return 'Kalteng 1', raw
    if 'KALTENG 2' in s_upper or 'KALIMANTAN TENGAH 2' in s_upper: return 'Kalteng 2', raw
    if 'KALTENG 3' in s_upper or 'KALIMANTAN TENGAH 3' in s_upper: return 'Kalteng 3', raw
    if 'BABEL' in s_upper or 'BANGKA' in s_upper: return 'Babel', raw
    if 'SUMBAR' in s_upper or 'SUMATERA BARAT' in s_upper: return 'Sumbar', raw
    if 'SUMSEL' in s_upper or 'SUMATERA SELATAN' in s_upper: return 'Sumsel', raw
    if 'SULTENG' in s_upper or 'SULAWESI TENGAH' in s_upper: return 'Sulteng', raw
    if 'SULTRA' in s_upper or 'SULAWESI TENGGARA' in s_upper: return 'Sultra', raw
    if 'PAPUA' in s_upper: return 'Papua Selatan', raw
    if 'ACEH' in s_upper: return 'Aceh', raw
    if 'JAMBI' in s_upper: return 'Jambi', raw
    if 'RIAU 1' in s_upper: return 'Riau 1', raw
    if 'RIAU 2' in s_upper: return 'Riau 2', raw
    if 'RIAU 3' in s_upper: return 'Riau 3', raw
    if 'RIAU 4' in s_upper: return 'Riau 4', raw

    return s.replace('Regional', '').strip(), raw

def normalize_cro_tambahan(cro_raw):
    if not cro_raw:
        return '-'
    s = str(cro_raw).strip().upper()
    if s == 'CRO I': return 'CRO 1'
    if s == 'CRO II': return 'CRO 2'
    if s == 'CRO III': return 'CRO 3'
    if s == 'CRO IV': return 'CRO 4'
    if s == 'CRO V': return 'CRO 5'
    if s == 'CRO VI': return 'CRO 6'
    if s == 'CRO VII': return 'CRO 7'
    if s == 'CRO VIII': return 'CRO 8'
    return s

def normalize_region_tambahan(r_raw):
    if not r_raw:
        return 'Aceh', 'Regional Aceh'
    s = str(r_raw).strip()
    raw = s
    s_upper = s.upper()

    if 'SUMUT 1' in s_upper or 'SUMATERA UTARA 1' in s_upper: return 'Sumut 1', raw
    if 'SUMUT 2' in s_upper or 'SUMATERA UTARA 2' in s_upper: return 'Sumut 2', raw
    if 'TORGANDA 2' in s_upper or 'RIAU 4' in s_upper: return 'Riau 4', raw
    if 'TORGANDA' in s_upper: return 'Sumut 2', raw
    if 'KALBAR 1' in s_upper or 'DUTA PALMA' in s_upper: return 'Kalbar 1', raw
    if 'KALBAR 2' in s_upper or 'KALIMANTAN BARAT 2' in s_upper: return 'Kalbar 2', raw
    if 'KALSEL 1' in s_upper or 'KALIMANTAN SELATAN 1' in s_upper: return 'Kalsel 1', raw
    if 'KALSEL 2' in s_upper or 'KALIMANTAN SELATAN 2' in s_upper: return 'Kalsel 2', raw
    if 'KALTARA' in s_upper or 'KALIMANTAN UTARA' in s_upper: return 'Kaltara', raw
    if 'KALTIM' in s_upper or 'KALIMANTAN TIMUR' in s_upper: return 'Kaltim', raw
    if 'KALTENG 1' in s_upper or 'KALIMANTAN TENGAH 1' in s_upper: return 'Kalteng 1', raw
    if 'KALTENG 2' in s_upper or 'KALIMANTAN TENGAH 2' in s_upper: return 'Kalteng 2', raw
    if 'KALTENG 3' in s_upper or 'KALIMANTAN TENGAH 3' in s_upper: return 'Kalteng 3', raw
    if 'BABEL' in s_upper or 'BANGKA BELITUNG' in s_upper or 'BANGKA' in s_upper: return 'Babel', raw
    if 'SUMBAR' in s_upper or 'SUMATERA BARAT' in s_upper: return 'Sumbar', raw
    if 'SUMSEL' in s_upper or 'SUMATERA SELATAN' in s_upper: return 'Sumsel', raw
    if 'SULTENG' in s_upper or 'SULAWESI TENGAH' in s_upper: return 'Sulteng', raw
    if 'SULTRA' in s_upper or 'SULAWESI TENGGARA' in s_upper: return 'Sultra', raw
    if 'PAPUA' in s_upper: return 'Papua Selatan', raw
    if 'ACEH' in s_upper: return 'Aceh', raw
    if 'JAMBI' in s_upper: return 'Jambi', raw
    if 'RIAU 1' in s_upper: return 'Riau 1', raw
    if 'RIAU 2' in s_upper: return 'Riau 2', raw
    if 'RIAU 3' in s_upper: return 'Riau 3', raw
    if 'RIAU 4' in s_upper: return 'Riau 4', raw

    return s.replace('Regional', '').strip(), raw

def clean_name(name):
    if not name:
        return ""
    s = str(name).strip().upper()
    s = re.sub(r'^(PT|CV|KOPERASI|KT|KUD|KELOMPOK TANI|UD)\s+', '', s)
    s = re.sub(r'\s+(TBK|LTD|PERSERO|CO|INC)$', '', s)
    s = re.sub(r'[^A-Z0-9\s]', '', s)
    s = re.sub(r'\s+', ' ', s).strip()
    return s

def to_num(val):
    if val is None or val == '-' or val == '':
        return 0
    try:
        return float(val)
    except:
        return 0

items = []
current_id = 1

# 1. Parse main Excel sheet
print(f"Reading main Excel file: {excel_path}")
wb = openpyxl.load_workbook(excel_path, data_only=True)
sheet = wb['DATA RAW']

for r in range(2, sheet.max_row + 1):
    cro = sheet.cell(r, 1).value
    wilayah = sheet.cell(r, 2).value
    nama_kebun = sheet.cell(r, 3).value

    if not cro and not nama_kebun:
        continue

    region, region_raw = normalize_region(wilayah)
    name_tag = sheet.cell(r, 4).value or f"KB-TAG-{current_id:03d}"
    luasan = to_num(sheet.cell(r, 5).value)
    req_tk = int(to_num(sheet.cell(r, 6).value))
    tk_mei = int(to_num(sheet.cell(r, 7).value))
    tk_juni = int(to_num(sheet.cell(r, 8).value))
    tk_juli = int(to_num(sheet.cell(r, 9).value))
    tk_agustus = int(to_num(sheet.cell(r, 10).value))
    target_juli = int(to_num(sheet.cell(r, 11).value))
    target_agustus = int(to_num(sheet.cell(r, 12).value))

    item = {
        "id": current_id,
        "cro": str(cro).strip() if cro else "-",
        "region_raw": region_raw,
        "region": region,
        "nama_kebun": str(nama_kebun).strip() if nama_kebun else "-",
        "name_tag": str(name_tag).strip(),
        "luasan": round(luasan, 2),
        "req_tk": req_tk,
        "tk_mei": tk_mei,
        "tk_juni": tk_juni,
        "target_juli": target_juli,
        "target_agustus": target_agustus,
        "target_september": 0,
        "tk_juli": tk_juli,
        "tk_agustus": tk_agustus,
        "updated_by": "EXCEL_SYNC",
        "updated_at": "2026-08-05T13:36:00.000Z"
    }
    items.append(item)
    current_id += 1

print(f"Parsed {len(items)} entries from DATA TK PANEN AGRIPAM.xlsx!")

# 2. Parse additional Excel sheet (Only adding 78 truly new non-overlapping ones)
if os.path.exists(excel_tambahan_path):
    print(f"Reading additional Excel file: {excel_tambahan_path}")
    wb_t = openpyxl.load_workbook(excel_tambahan_path, data_only=True)
    sheet_t = wb_t.active
    
    # Collect existing names from main list
    existing_names = [item['nama_kebun'].strip().upper() for item in items]
    existing_clean_names = [clean_name(item['nama_kebun']) for item in items]
    added_names = set()
    
    added_count = 0
    skipped_count = 0
    
    for r in range(2, sheet_t.max_row + 1):
        cro_raw = sheet_t.cell(r, 1).value
        wilayah_raw = sheet_t.cell(r, 2).value
        nama_kebun = sheet_t.cell(r, 4).value
        
        if not wilayah_raw and not nama_kebun:
            continue
            
        nama_kebun_str = str(nama_kebun).strip() if nama_kebun else ""
        if not nama_kebun_str:
            continue
            
        clean_raw = clean_name(nama_kebun_str)
        
        is_match = False
        # Check exact
        if any(nama_kebun_str.upper() == n for n in existing_names):
            is_match = True
        
        # Check fuzzy/substring
        if not is_match and len(clean_raw) > 3:
            if any(len(cn) > 3 and (clean_raw in cn or cn in clean_raw) for cn in existing_clean_names):
                is_match = True
                
        if is_match or nama_kebun_str.upper() in added_names:
            skipped_count += 1
            continue
            
        added_names.add(nama_kebun_str.upper())
        cro = normalize_cro_tambahan(cro_raw)
        region, region_raw = normalize_region_tambahan(wilayah_raw)
        
        item = {
            "id": current_id,
            "cro": cro,
            "region_raw": region_raw,
            "region": region,
            "nama_kebun": nama_kebun_str,
            "name_tag": "",
            "luasan": 0,
            "req_tk": 0,
            "tk_mei": 0,
            "tk_juni": 0,
            "target_juli": 0,
            "target_agustus": 0,
            "target_september": 0,
            "tk_juli": 0,
            "tk_agustus": 0,
            "updated_by": "EXCEL_SYNC_TAMBAHAN",
            "updated_at": "2026-08-18T10:00:00.000Z"
        }
        items.append(item)
        current_id += 1
        added_count += 1
        
    print(f"Parsed and appended {added_count} new entries from DATA TAMBAHAN PENUGASAN APN.xlsx (Skipped {skipped_count} duplicates/matches)!")

# Write data_kebun_tk.json
with open(json_path, 'w', encoding='utf-8') as f:
    json.dump(items, f, indent=2)

print(f"Saved {json_path}!")

# Generate SQL File
def clean_sql_str(val):
    if val is None:
        return 'NULL'
    s = str(val).replace("'", "''").strip()
    return f"'{s}'"

sql_lines = []
sql_lines.append("-- ============================================================================")
sql_lines.append("-- AGRIPAM - SUPABASE SQL EDITOR SCRIPT: DATA KEBUN TK PANEN")
sql_lines.append(f"-- Source Data: Combined ({len(items)} Entri Kebun)")
sql_lines.append("-- Cara Penggunaan:")
sql_lines.append("-- 1. Buka Supabase Dashboard Project Anda -> Menu 'SQL Editor'")
sql_lines.append("-- 2. Copy (Salin) seluruh isi kode SQL ini dan Paste di SQL Editor")
sql_lines.append("-- 3. Klik tombol 'Run' (Jalankan)")
sql_lines.append("-- ============================================================================\n")

sql_lines.append("-- 1. HAPUS TABEL LAMA JIKA ADA")
sql_lines.append("DROP TABLE IF EXISTS data_kebun_tk CASCADE;\n")

sql_lines.append("-- 2. BUAT STRUKTUR TABEL data_kebun_tk")
sql_lines.append("""CREATE TABLE data_kebun_tk (
    id INT PRIMARY KEY,
    cro VARCHAR(50),
    region VARCHAR(100) NOT NULL,
    region_raw VARCHAR(100),
    nama_kebun VARCHAR(255) NOT NULL,
    name_tag VARCHAR(100),
    luasan NUMERIC(10, 2) DEFAULT 0,
    req_tk INT DEFAULT 0,
    tk_mei INT DEFAULT 0,
    tk_juni INT DEFAULT 0,
    target_juli INT DEFAULT 0,
    target_agustus INT DEFAULT 0,
    target_september INT DEFAULT 0,
    tk_juli INT DEFAULT 0,
    tk_agustus INT DEFAULT 0,
    updated_by VARCHAR(100) DEFAULT 'SYSTEM',
    updated_at TIMESTAMP WITH TIME ZONE DEFAULT CURRENT_TIMESTAMP,
    created_at TIMESTAMP WITH TIME ZONE DEFAULT CURRENT_TIMESTAMP
);\n""")

sql_lines.append("-- 3. BUAT INDEX UNTUK PERFORMA QUERY CEPAT")
sql_lines.append("CREATE INDEX idx_data_kebun_tk_region ON data_kebun_tk(region);")
sql_lines.append("CREATE INDEX idx_data_kebun_tk_cro ON data_kebun_tk(cro);")
sql_lines.append("CREATE INDEX idx_data_kebun_tk_name_tag ON data_kebun_tk(name_tag);\n")

sql_lines.append("-- 4. AKTIFKAN ROW LEVEL SECURITY (RLS) & KEBIJAKAN AKSES PUBLIC")
sql_lines.append("ALTER TABLE data_kebun_tk ENABLE ROW LEVEL SECURITY;")
sql_lines.append("CREATE POLICY \"Allow public select on data_kebun_tk\" ON data_kebun_tk FOR SELECT USING (true);")
sql_lines.append("CREATE POLICY \"Allow public insert on data_kebun_tk\" ON data_kebun_tk FOR INSERT WITH CHECK (true);")
sql_lines.append("CREATE POLICY \"Allow public update on data_kebun_tk\" ON data_kebun_tk FOR UPDATE USING (true);\n")

sql_lines.append(f"-- 5. INSERT {len(items)} DATA KEBUN")

chunk_size = 100
for i in range(0, len(items), chunk_size):
    chunk = items[i:i + chunk_size]
    sql_lines.append(f"-- Batch Insert Data Kebun ({i + 1} s/d {i + len(chunk)})")
    sql_lines.append("INSERT INTO data_kebun_tk (id, cro, region, region_raw, nama_kebun, name_tag, luasan, req_tk, tk_mei, tk_juni, target_juli, target_agustus, target_september, tk_juli, tk_agustus, updated_by) VALUES")
    
    values_list = []
    for item in chunk:
        item_id = item.get('id', 0)
        cro = clean_sql_str(item.get('cro'))
        region = clean_sql_str(item.get('region'))
        region_raw = clean_sql_str(item.get('region_raw'))
        nama_kebun = clean_sql_str(item.get('nama_kebun'))
        name_tag = clean_sql_str(item.get('name_tag'))
        luasan = item.get('luasan', 0)
        req_tk = item.get('req_tk', 0)
        tk_mei = item.get('tk_mei', 0)
        tk_juni = item.get('tk_juni', 0)
        target_juli = item.get('target_juli', 0)
        target_agustus = item.get('target_agustus', 0)
        target_september = item.get('target_september', 0)
        tk_juli = item.get('tk_juli', 0)
        tk_agustus = item.get('tk_agustus', 0)
        updated_by = clean_sql_str(item.get('updated_by', 'SYSTEM'))

        val_str = f"({item_id}, {cro}, {region}, {region_raw}, {nama_kebun}, {name_tag}, {luasan}, {req_tk}, {tk_mei}, {tk_juni}, {target_juli}, {target_agustus}, {target_september}, {tk_juli}, {tk_agustus}, {updated_by})"
        values_list.append(val_str)

    sql_lines.append(",\n".join(values_list) + ";\n")

sql_lines.append("-- ============================================================================")
sql_lines.append(f"-- SELESAI: {len(items)} Data Kebun Berhasil Dibuat dan Di-seed ke Supabase!")
sql_lines.append("-- ============================================================================")

sql_content = "\n".join(sql_lines)

with open(sql_path, 'w', encoding='utf-8') as f:
    f.write(sql_content)

with open(sql_root_path, 'w', encoding='utf-8') as f:
    f.write(sql_content)

print(f"SUCCESSFULLY UPDATED BOTH SQL FILES:\n - {sql_path}\n - {sql_root_path}")
